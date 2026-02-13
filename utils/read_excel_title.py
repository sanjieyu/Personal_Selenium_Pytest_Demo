# Author:Yi Sun(Tim) 2020-12-12

'''Read Excel function by using Openpyxl'''
'''
openpyxl does not support the old .xls file format, 
please use xlrd to read this file, or convert it to the more recent .xlsx file format
'''
'''
this function is print the first line title,
then print each value from the second line to match the columu title of the first line.IT'S a DICT, like this:
total row is: 6
total columns is: 3
title is: ['id', 'username', 'pwd']
all data is: [{'id': 1, 'username': None, 'pwd': None}, {'id': 2, 'username': 'sanjieyu', 'pwd': 123456}, {'id': 3, 
'username': 'preet.kaur', 'pwd': 123456}, {'id': 4, 'username': 'preet.kaur', 'pwd': 'Preet@123'}, {'id': 5, 'username':
 'timnew', 'pwd': 'Tims@123'}]
'''

from openpyxl import load_workbook
import os

class ExcelData():
    def __init__(self):
        current_loc = os.path.abspath("..")
        excel_path = os.path.join(current_loc,'Config\\api_EGD_Dev.xlsx')
        self.wb = load_workbook(excel_path)
        self.sheet = self.wb["Sheet1"]
        self.rows = self.sheet.max_row
        self.columns = self.sheet.max_column

    def read_excel(self):
        all_data = []
        title = [self.sheet.cell(1, column).value for column in range(1, self.columns + 1)]
        for row in range(2,self.rows + 1):
            data = [self.sheet.cell(row, column).value for column in range(1, self.columns + 1)]
            all_data.append(dict(zip(title,data)))
        # print('all data is:',all_data[-1])  # print the last row data
        # print('all data is:', all_data)  # print all data
        return all_data

if __name__ == "__main__":
    readexcel = ExcelData()
    readexcel.read_excel()